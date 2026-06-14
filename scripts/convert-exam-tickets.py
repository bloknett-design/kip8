#!/usr/bin/env python3
"""
Скачивает xlsx-файл с OneDrive и конвертирует его в JSON
для использования в PWA «КИПиА».

Запускается автоматически через GitHub Actions (workflow_dispatch или schedule),
либо вручную.
"""

import json
import os
import re
import sys

import openpyxl
import requests

# ============================================================
# Конфигурация — ссылка на общий доступ к файлу OneDrive
# ============================================================
ONEDRIVE_SHARE_URL = os.environ.get(
    "ONEDRIVE_SHARE_URL",
    "https://1drv.ms/x/c/c9414adb26fe5b28/IQDqsaT9uFR_T6HBWRyYhhEtAXDVcDPDoT3-O_um4E6V7O0?e=IKtgRl",
)

# Путь к выходному JSON-файлу (относительно корня репозитория)
OUTPUT_PATH = os.environ.get("OUTPUT_PATH", "data/exam-tickets.json")

# Маппинг листов → идентификаторы и заголовки
SHEETS_CONFIG = {
    "4 разряд": {"id": "tickets-4", "title": "Билеты на 4 разряд"},
    "5 разряд": {"id": "tickets-5", "title": "Билеты на 5 разряд"},
    "6 разряд": {"id": "tickets-6", "title": "Билеты на 6 разряд"},
    "До 1000 В": {"id": "tickets-1000v", "title": "Билеты до 1000 В"},
}


def download_xlsx(share_url: str, dest: str) -> bool:
    """Скачивает xlsx-файл по ссылке общего доступа OneDrive."""
    import base64

    # Способ 1: Microsoft Graph API shares (для публичных ссылок)
    encoded = base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    api_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded}/root/content"
    r = requests.get(api_url, allow_redirects=True, timeout=30)
    if r.status_code == 200 and len(r.content) > 100 and r.content[:2] == b"PK":
        with open(dest, "wb") as f:
            f.write(r.content)
        print(f"Скачано через Graph API: {len(r.content)} байт")
        return True

    # Способ 2: Через web-reader / извлечение download URL из страницы
    r = requests.get(share_url, allow_redirects=True, timeout=20)
    page_text = r.text

    # Ищем tempauth download URL
    auth_match = re.search(
        r"(https://[^\s\"\\]+download\.aspx[^\s\"\\]*tempauth=[^\s\"\\]+)",
        page_text,
    )
    if auth_match:
        dl_url = (
            auth_match.group(1)
            .replace("\\u0026", "&")
            .replace("&amp;", "&")
        )
        r2 = requests.get(dl_url, allow_redirects=True, timeout=30)
        if r2.status_code == 200 and r2.content[:2] == b"PK":
            with open(dest, "wb") as f:
                f.write(r2.content)
            print(f"Скачано через tempauth URL: {len(r2.content)} байт")
            return True

    # Способ 3: Прямой download URL без tempauth
    simple_match = re.search(
        r"(https://onedrive\.live\.com/[^\s\"\\]+download\.aspx[^\s\"\\]*)",
        page_text,
    )
    if simple_match:
        dl_url = (
            simple_match.group(1)
            .replace("\\u0026", "&")
            .replace("&amp;", "&")
        )
        r3 = requests.get(dl_url, allow_redirects=True, timeout=30)
        if r3.status_code == 200 and r3.content[:2] == b"PK":
            with open(dest, "wb") as f:
                f.write(r3.content)
            print(f"Скачано через простой download URL: {len(r3.content)} байт")
            return True

    print("Не удалось скачать файл ни одним из способов", file=sys.stderr)
    return False


def convert_xlsx_to_json(xlsx_path: str, json_path: str) -> bool:
    """Конвертирует xlsx-файл в JSON."""
    wb = openpyxl.load_workbook(xlsx_path)
    all_data = {}

    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"Лист «{sheet_name}» не найден, пропускаю", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else "" for h in rows[0]]
        data_rows = []

        for row in rows[1:]:
            obj = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    obj[key] = str(val) if val is not None else ""
            data_rows.append(obj)

        all_data[config["id"]] = {
            "title": config["title"],
            "sheet": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "total": len(data_rows),
        }
        print(f"  {sheet_name}: {len(data_rows)} строк")

    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    size = os.path.getsize(json_path)
    print(f"JSON сохранён: {json_path} ({size / 1024:.1f} КБ)")
    return True


def main():
    xlsx_dest = "/tmp/exam_tickets.xlsx"

    print("Скачивание xlsx с OneDrive...")
    if not download_xlsx(ONEDRIVE_SHARE_URL, xlsx_dest):
        sys.exit(1)

    print("Конвертация xlsx → JSON...")
    if not convert_xlsx_to_json(xlsx_dest, OUTPUT_PATH):
        sys.exit(1)

    print("Готово!")


if __name__ == "__main__":
    main()
